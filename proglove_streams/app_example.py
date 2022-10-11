"""Minimal example dumping whatever event it receives."""
from asyncio.windows_events import NULL
from pickle import FALSE
import time
#import logging
import argparse
from tkinter.font import NORMAL
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import DISABLED, NS, VERTICAL, Scrollbar, ttk, Entry
from tkinter.filedialog import askopenfile
import serial.tools.list_ports
from datetime import datetime
import os.path

from proglove_streams.logging import init_logging
from proglove_streams.client import Client
from proglove_streams.gateway import Gateway, GatewayMessageHandler
from proglove_streams.exception import ProgloveStreamsException
from proglove_streams.models.scan import ScanEvent
from proglove_streams.models.scanner_state import ScannerStateEvent
from proglove_streams.models.error import ErrorEvent
from proglove_streams.models.gateway_state import GatewayStateEvent
from proglove_streams.models.button_pressed import ButtonPressedEvent


root = tk.Tk()
root.resizable(False, False)
root.title("MARK Display Excel Lookup")
if os.path.exists('MARKicon.ico'):
    root.iconbitmap("MARKicon.ico")

def text_box_log(str):
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    text_box.configure(state=NORMAL)
    text_box.insert('end', "["+current_time+"] "+str+"\n")
    text_box.see("end")
    text_box.configure(state=DISABLED)

def mirror_PG3(title1, value1, title2, value2, title3, value3):
    text_box.configure(state=NORMAL)
    text_box.tag_config("title", font=("Roboto bold", 10))
    text_box.tag_config("value", font=("Roboto bold", 12))
    text_box.insert('end', title1+"\n", "title")
    text_box.insert('end', value1+"\n", "value")
    text_box.insert('end', title2+"\n", "title")
    text_box.insert('end', value2+"\n", "value")
    text_box.insert('end', title3+"\n", "title")
    text_box.insert('end', value3+"\n", "value")
    text_box.see("end")
    text_box.configure(state=DISABLED)

def find_row_number(event: ScanEvent):
    for row in range(1, currentSheet.max_row + 1):
        for column in "A":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            #text_box_log(currentSheet[cell_name].value)
            if currentSheet[cell_name].value == event.scan_code:
                text_box_log("Cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
                return row

def _lookup_and_display(client: Gateway, event: ScanEvent):
    for sheet in allSheetNames:
        #text_box_log("Current sheet name is {}" .format(sheet))
        currentSheet = theFile[sheet]
        rowNumber = (find_row_number(event))
        #text_box_log("rowNumber = " + repr(rowNumber))

    #display data if we found the barcode
    if rowNumber:
        title_cell1 = "{}{}".format("A", 1)
        title_cell2 = "{}{}".format("B", 1)
        title_cell3 = "{}{}".format("C", 1)
        title_cell4 = "{}{}".format("D", 1)

        value_cell1 = "{}{}".format("A", rowNumber)
        value_cell2 = "{}{}".format("B", rowNumber)
        value_cell3 = "{}{}".format("C", rowNumber)
        value_cell4 = "{}{}".format("D", rowNumber)

        title_value1 = currentSheet[title_cell1].value
        title_value2 = currentSheet[title_cell2].value
        title_value3 = currentSheet[title_cell3].value + "\\" + currentSheet[title_cell4].value

        value_string1 = currentSheet[value_cell1].value
        value_string2 = currentSheet[value_cell2].value
        value_string3 = currentSheet[value_cell3].value + "\\" + repr(currentSheet[value_cell4].value)

        text_box_log("MARK Display shows:")
        mirror_PG3(title_value1, value_string1, title_value2, value_string2, title_value3, value_string3)

        client.set_display(str(event.device_serial), 'PG3',
                        display_fields=[
                            {
                                "display_field_id": 1,
                                "display_field_header": title_value1,
                                "display_field_text": value_string1
                            },
                            {
                                "display_field_id": 2,
                                "display_field_header": title_value2,
                                "display_field_text": value_string2
                            },
                            {
                                "display_field_id": 3,
                                "display_field_header": title_value3,
                                "display_field_text": value_string3
                            }
                        ])
    #display error message if barcode is not found                    
    else:
        client.set_display(str(event.device_serial), 'PG1E',
                        display_fields=[
                            {
                                "display_field_id": 1,
                                "display_field_header": "",
                                "display_field_text": event.scan_code + "\nNot Found!"
                            }
                        ])
        text_box_log(event.scan_code + " not found in database!")


def _set_display(client: Gateway, event: ScanEvent):
    client.set_display(str(event.device_serial), 'PG3',
                       display_fields=[
                           {
                               "display_field_id": 1,
                               "display_field_header": "Storage Unit",
                               "display_field_text": "R15"
                           },
                           {
                               "display_field_id": 2,
                               "display_field_header": "Item",
                               "display_field_text": "Engine 12"
                           },
                           {
                               "display_field_id": 3,
                               "display_field_header": "Quantity",
                               "display_field_text": "10"
                           }
                       ])


def _block_trigger(client: Gateway, event: ScanEvent):
    client.set_trigger_block(str(event.device_serial), True,
                             ["TRIGGER_SINGLE_CLICK"], [],
                             time_validity_duration=3000)


def _unblock_trigger(client: Gateway, event: ScanEvent):
    client.set_trigger_block(str(event.device_serial), False,
                             [], [])


def on_connected(_client: Client, event: ScannerStateEvent) -> None:
    """On connected event callback."""
    text_box_log("device connected: " + event.device_serial)


def on_disconnected(_client: Client, event: ScannerStateEvent) -> None:
    """On disconnected event callback."""
    text_box_log("device disconnected: " + event.device_serial)


def on_scan(client: Client, event: ScanEvent) -> None:
    """On scan event callback."""
    if not isinstance(client, Gateway):
        return

    text_box_log("Scan received: device " + event.device_serial + " Data: " + repr(event.scan_code))

    scan_code = str(event.scan_code).split('\r')[0]

    if scan_code == 'DISPLAY':
        _set_display(client, event)
    elif scan_code == 'BLOCK':
        _block_trigger(client, event)
    elif scan_code == 'UNBLOCK':
        _unblock_trigger(client, event)
    elif scan_code == 'FEEDBACK_OK':
        client.send_feedback(str(event.device_serial), 'FEEDBACK_POSITIVE')
    elif scan_code == 'FEEDBACK_NOK':
        client.send_feedback(str(event.device_serial), 'FEEDBACK_NEGATIVE')
    elif scan_code == 'STATE':
        client.get_gateway_state()
    else:
        _lookup_and_display(client, event)
        


def on_error(_client: Client, event: ErrorEvent) -> None:
    """On error event callback."""
    text_box_log("Error received: "+ event.error_code)

### These callback functions are from the demo code, but not used for any functionality in this app###
### START ###
def on_gateway_state_event(_client: Client, event: GatewayStateEvent):
    """On Gateway state event callback."""
    # logger.info('''Gateway state received: serial: %s version: %s
    #                connected devices: %s''',
    #             event.gateway_serial,
    #             event.gateway_app_version,
    #             ','.join([d.device_serial
    #                       for d in event.device_connected_list]))


def on_button_pressed_event(_client: Client,
                            event: ButtonPressedEvent) -> None:
    """On error event callback."""
    # logger.info('button pressed: device %s, trigger gesture: %s',
    #             event.device_serial,
    #             event.trigger_gesture)
### END ###

def open_file(root, directory):
    file = askopenfile(parent=root, mode='rb', title="Choose a file", filetype=[("Excel File", "*.xls; *.xlsx")])
    if file:
        directory.set(file.name)

def connect_btn_clicked(button, port, path):
    handler = GatewayMessageHandler(
        on_scanner_connected=on_connected,
        on_scanner_disconnected=on_disconnected,
        on_scan=on_scan,
        on_error=on_error,
        on_gateway_state_event=on_gateway_state_event,
        on_button_pressed=on_button_pressed_event
    )

    try:
        global gateway
        gateway = Gateway(handler, port, 115200)
        gateway.start()
    except ProgloveStreamsException as e:
        text_box_log("StreamsAPI Exception: " + e)
        return

    text_box_log("Connection to "+ port + " Successful!")
        
    button['state'] = DISABLED

    global theFile
    global allSheetNames
    global currentSheet
    try:
        theFile = load_workbook(path)
        allSheetNames = theFile.sheetnames
        if allSheetNames:
            currentSheet = theFile[allSheetNames[0]]
            text_box_log("Opened workbook "+path+" to use as database")
    except:
        gateway.stop()
        button['state'] = NORMAL
        text_box_log("Could not open spreadsheet path!")

    

    root.protocol("WM_DELETE_WINDOW", lambda:on_close(root, gateway))

def on_close(root, gateway):
    if gateway:
        gateway.stop()
    root.destroy()
    exit(0)

def app_example():
#GUI definitions
#COM port dropdown
    ports = serial.tools.list_ports.comports()
    ports_list = []
    for p in ports:
        ports_list.append(p.device)

    com_label = ttk.Label(text="COM Port:")
    com_label.grid(column=0, row=0, padx = 5, pady=5)

    selected_com_port = tk.StringVar()
    comport_cb = ttk.Combobox(root, textvariable=selected_com_port)
    comport_cb['values'] = ports_list
    comport_cb['state'] = 'readonly'
    comport_cb.grid(column=1, row=0, padx = 5, pady=5)

#selected path for spreadsheet
    path_label = ttk.Label(text="Spreadsheet Path:")
    path_label.grid(column=2, row=0)

    file_path = tk.StringVar()
    file_path_input = Entry(root, textvariable=file_path, width=30)
    file_path_input.grid(column=3, row=0, padx = 5, pady=5)

#browse button
    browse_text = tk.StringVar()
    browse_btn = tk.Button(root, textvariable=browse_text, command=lambda:open_file(root, file_path))
    browse_text.set("Browse")
    browse_btn.grid(column=4, row=0, padx = 5, pady=5)

#Connect button
    connect_text = tk.StringVar()
    connect_btn = tk.Button(root, textvariable=connect_text, font = ('Roboto', 12, 'bold'), bg="#f47920", fg="white", height = 2, width = 8, command=lambda:connect_btn_clicked(connect_btn, selected_com_port.get(), file_path.get()))
    connect_text.set("Connect")
    connect_btn.grid(column=1, row=1, columnspan = 3, pady=5)

#text box for console
    global text_box
    text_box = tk.Text(root, height=10, width=63)
    text_box.configure(state=DISABLED)
    text_box.grid(column=0, row=2, columnspan=5, pady=15)

    sb = tk.Scrollbar(root, orient=VERTICAL)
    sb.grid(column=4, row=2, sticky=NS, padx=15, pady=15)
    text_box.config(yscrollcommand=sb.set)
    sb.config(command=text_box.yview)

    device = selected_com_port
    baudrate = 115200
    
    root.protocol("WM_DELETE_WINDOW", lambda:on_close(root, NULL))
    root.mainloop()



